#!/usr/bin/env python3
"""
Sistema de procesamiento de archivos Excel para Picking Center.

Optimized for serverless: reads data with openpyxl read_only mode,
then modifies the xlsx directly at the ZIP/XML level to avoid the
slow full load_workbook cycle.
"""

import sys
import os
import re
import glob
import shutil
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import time as dt_time

import openpyxl

# xlsx XML namespaces
NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
NS_REL = 'http://schemas.openxmlformats.org/package/2006/relationships'
NS_CT = 'http://schemas.openxmlformats.org/package/2006/content-types'

ET.register_namespace('', NS)
ET.register_namespace('r', NS_R)


def detectar_archivos(ruta_carpeta):
    archivos_xlsx = glob.glob(os.path.join(ruta_carpeta, "*.xlsx"))
    archivos_xlsx = [f for f in archivos_xlsx if not os.path.basename(f).startswith("~$")]
    if len(archivos_xlsx) < 2:
        raise Exception("Se necesitan al menos 2 archivos .xlsx en la carpeta.")
    archivo1 = archivo2 = None
    for f in archivos_xlsx:
        nombre = os.path.basename(f).upper()
        if "RESULTADO" in nombre or "OUTPUT" in nombre:
            continue
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            hojas = wb.sheetnames
            wb.close()
            if any("PROGRAMACI" in h.upper() for h in hojas):
                if any("PLANIFICACION" in h for h in hojas):
                    archivo1 = f
                elif "RUTA" in hojas:
                    archivo2 = f
                elif archivo1 is None:
                    archivo1 = f
            elif "RUTA" in hojas:
                archivo2 = f
        except Exception:
            continue
    if archivo1 is None or archivo2 is None:
        raise Exception("No se pudieron identificar los archivos automáticamente.")
    return archivo1, archivo2


def extraer_mapeo_archivo1(ruta_archivo1):
    wb = openpyxl.load_workbook(ruta_archivo1, read_only=True, data_only=True)
    hoja_prog = None
    for nombre in wb.sheetnames:
        if "PROGRAMACI" in nombre.upper():
            hoja_prog = wb[nombre]
            break
    if hoja_prog is None:
        wb.close()
        raise Exception("No se encontró la hoja PROGRAMACIÓN en el Archivo 1.")

    rows = list(hoja_prog.iter_rows(values_only=True))
    wb.close()

    header_row_idx = 0
    for i, row in enumerate(rows[:10]):
        if any(v and "TRANSPORTE" in str(v).upper() for v in row if v):
            header_row_idx = i
            break

    header = rows[header_row_idx]
    col_map = {}
    for j, val in enumerate(header):
        if val:
            v = str(val).upper().strip()
            if "TRANSPORTE" in v:
                col_map["n_transporte"] = j
            elif "PLACA" in v:
                col_map["placa"] = j
            elif "CARGO" in v:
                col_map["cargo"] = j

    if "n_transporte" not in col_map or "placa" not in col_map:
        raise Exception(f"Columnas necesarias no encontradas. Encontradas: {col_map}")

    mapeo = {}
    for row in rows[header_row_idx + 1:]:
        n_trans = row[col_map["n_transporte"]] if len(row) > col_map["n_transporte"] else None
        placa = row[col_map["placa"]] if len(row) > col_map["placa"] else None
        cargo = row[col_map["cargo"]] if "cargo" in col_map and len(row) > col_map["cargo"] else "Conductor"
        if n_trans and placa and str(cargo).strip().lower() == "conductor":
            mapeo[str(n_trans).strip()] = str(placa).strip()
    return mapeo


def derivar_opl(placa_generica):
    if not placa_generica:
        return ""
    return re.sub(r'\d+$', '', str(placa_generica).strip())


def col_letter(col_num):
    """1-based column number to letter (1=A, 17=Q, etc.)"""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def get_sheet_file_map(z):
    """Returns dict mapping sheet name -> sheet XML path inside xlsx."""
    wb_xml = ET.fromstring(z.read('xl/workbook.xml'))
    rels_xml = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))

    rid_to_name = {}
    for sheet in wb_xml.iter(f'{{{NS}}}sheet'):
        rid_to_name[sheet.get(f'{{{NS_R}}}id')] = sheet.get('name')

    name_to_file = {}
    for rel in rels_xml.iter():
        rid = rel.get('Id')
        target = rel.get('Target')
        if rid and target and rid in rid_to_name:
            name_to_file[rid_to_name[rid]] = f'xl/{target}'
    return name_to_file


def modify_sheet_xml(sheet_xml_bytes, cell_updates):
    """
    Modify cells in a sheet XML using string manipulation to preserve
    original namespaces exactly. cell_updates: {(row, col): value}
    """
    xml_str = sheet_xml_bytes.decode('utf-8')

    # Group updates by row
    updates_by_row = defaultdict(dict)
    for (row, col), value in cell_updates.items():
        updates_by_row[row][col] = value

    for row_num in sorted(updates_by_row.keys()):
        for col_num, value in updates_by_row[row_num].items():
            ref = f'{col_letter(col_num)}{row_num}'

            if value is None:
                continue

            # Build the cell XML
            if isinstance(value, (int, float)):
                cell_xml = f'<c r="{ref}"><v>{value}</v></c>'
            else:
                # Escape XML special chars
                escaped = str(value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                cell_xml = f'<c r="{ref}" t="inlineStr"><is><t>{escaped}</t></is></c>'

            # Check if cell already exists in the row
            cell_pattern = rf'<c\s+r="{ref}"[^/]*(?:/>|>.*?</c>)'
            if re.search(cell_pattern, xml_str, re.DOTALL):
                # Replace existing cell
                xml_str = re.sub(cell_pattern, cell_xml, xml_str, count=1, flags=re.DOTALL)
            else:
                # Find the row and append the cell before </row>
                row_pattern = rf'(<row[^>]*\s+r="{row_num}"[^>]*>)(.*?)(</row>)'
                row_match = re.search(row_pattern, xml_str, re.DOTALL)
                if row_match:
                    xml_str = xml_str[:row_match.end(2)] + cell_xml + xml_str[row_match.end(2):]
                else:
                    # Row doesn't exist - insert in correct sorted position
                    new_row = f'<row r="{row_num}">{cell_xml}</row>'
                    # Find the first row with r > row_num to insert before it
                    inserted = False
                    for m in re.finditer(r'<row\b[^>]*\s+r="(\d+)"', xml_str):
                        if int(m.group(1)) > row_num:
                            xml_str = xml_str[:m.start()] + new_row + xml_str[m.start():]
                            inserted = True
                            break
                    if not inserted:
                        xml_str = xml_str.replace('</sheetData>', new_row + '</sheetData>')

    return xml_str.encode('utf-8')


def build_new_sheet_xml(rows_data):
    """
    Build a complete sheet XML from scratch.
    rows_data: list of (row_num, [(col_num, value, bold), ...])
    """
    root = ET.Element(f'{{{NS}}}worksheet')

    # Column widths
    cols = ET.SubElement(root, f'{{{NS}}}cols')
    for i in range(1, 9):
        col_el = ET.SubElement(cols, f'{{{NS}}}col')
        col_el.set('min', str(i))
        col_el.set('max', str(i))
        col_el.set('width', '22')
        col_el.set('customWidth', '1')

    sheet_data = ET.SubElement(root, f'{{{NS}}}sheetData')

    for row_num, cells in rows_data:
        row_el = ET.SubElement(sheet_data, f'{{{NS}}}row')
        row_el.set('r', str(row_num))

        for col_num, value, bold in cells:
            ref = f'{col_letter(col_num)}{row_num}'
            c_el = ET.SubElement(row_el, f'{{{NS}}}c')
            c_el.set('r', ref)

            if value is None:
                continue
            elif isinstance(value, (int, float)):
                v_el = ET.SubElement(c_el, f'{{{NS}}}v')
                v_el.text = str(value)
            elif isinstance(value, dt_time):
                # Write as numeric time value with style referencing numFmtId 20 (h:mm)
                c_el.set('s', '4')  # style index 4 = numFmtId 21 (h:mm:ss) in this workbook
                v_el = ET.SubElement(c_el, f'{{{NS}}}v')
                v_el.text = str((value.hour * 3600 + value.minute * 60 + value.second) / 86400)
            else:
                c_el.set('t', 'inlineStr')
                is_el = ET.SubElement(c_el, f'{{{NS}}}is')
                t_el = ET.SubElement(is_el, f'{{{NS}}}t')
                t_el.text = str(value)

    return ET.tostring(root, xml_declaration=True, encoding='UTF-8')


def add_sheets_to_workbook(z_in, z_out, new_sheets, insert_before_sheet="RUTA"):
    """
    Copy all files from z_in to z_out, adding new sheets.
    Uses string manipulation instead of ElementTree to preserve
    original XML namespaces exactly as Excel expects them.
    new_sheets: list of (sheet_name, sheet_xml_bytes)
    """
    with zipfile.ZipFile(z_in, 'r') as zr:
        all_files = zr.namelist()

        # Find highest sheet number
        max_sheet_num = 0
        for fname in all_files:
            if fname.startswith('xl/worksheets/sheet') and fname.endswith('.xml'):
                try:
                    num = int(fname.replace('xl/worksheets/sheet', '').replace('.xml', ''))
                    max_sheet_num = max(max_sheet_num, num)
                except ValueError:
                    pass

        # Find highest rId and sheetId from workbook XML
        wb_text = zr.read('xl/workbook.xml').decode('utf-8')
        rels_text = zr.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        ct_text = zr.read('[Content_Types].xml').decode('utf-8')

        max_rid = 0
        for m in re.finditer(r'rId(\d+)', rels_text):
            max_rid = max(max_rid, int(m.group(1)))

        max_sheet_id = 0
        for m in re.finditer(r'sheetId="(\d+)"', wb_text):
            max_sheet_id = max(max_sheet_id, int(m.group(1)))

        # Prepare new sheet info
        new_sheet_info = []
        for i, (name, xml_bytes) in enumerate(new_sheets):
            sheet_num = max_sheet_num + 1 + i
            rid = f'rId{max_rid + 1 + i}'
            sheet_id = max_sheet_id + 1 + i
            fname = f'xl/worksheets/sheet{sheet_num}.xml'
            new_sheet_info.append((name, rid, sheet_id, fname, xml_bytes))

        # Remove existing Hoja3/Hoja8 entries from workbook.xml
        for name, *_ in new_sheet_info:
            wb_text = re.sub(rf'<sheet[^>]*name="{name}"[^/]*/>', '', wb_text)

        # Insert new sheet entries before the target sheet
        insert_tag = f'name="{insert_before_sheet}"'
        new_sheet_tags = ''
        for name, rid, sheet_id, fname, xml_bytes in new_sheet_info:
            new_sheet_tags += f'<sheet name="{name}" sheetId="{sheet_id}" r:id="{rid}"/>'

        # Insert before the target sheet
        insert_pos = wb_text.find(insert_tag)
        if insert_pos >= 0:
            # Find the start of the <sheet tag
            tag_start = wb_text.rfind('<sheet', 0, insert_pos)
            wb_text = wb_text[:tag_start] + new_sheet_tags + wb_text[tag_start:]
        else:
            # Fallback: insert before </sheets>
            wb_text = wb_text.replace('</sheets>', new_sheet_tags + '</sheets>')

        modified_wb = wb_text.encode('utf-8')

        # Add new relationships to workbook.xml.rels
        new_rels = ''
        for name, rid, sheet_id, fname, xml_bytes in new_sheet_info:
            target = fname.replace('xl/', '')
            new_rels += f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="{target}"/>'

        rels_text = rels_text.replace('</Relationships>', new_rels + '</Relationships>')
        modified_rels = rels_text.encode('utf-8')

        # Add content types for new sheets
        new_overrides = ''
        for name, rid, sheet_id, fname, xml_bytes in new_sheet_info:
            new_overrides += f'<Override PartName="/{fname}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'

        ct_text = ct_text.replace('</Types>', new_overrides + '</Types>')
        modified_ct = ct_text.encode('utf-8')

        # Write everything to output
        with zipfile.ZipFile(z_out, 'w', zipfile.ZIP_DEFLATED) as zw:
            for item in zr.infolist():
                if item.filename == 'xl/workbook.xml':
                    zw.writestr(item, modified_wb)
                elif item.filename == 'xl/_rels/workbook.xml.rels':
                    zw.writestr(item, modified_rels)
                elif item.filename == '[Content_Types].xml':
                    zw.writestr(item, modified_ct)
                else:
                    zw.writestr(item, zr.read(item.filename))

            # Add new sheet files
            for name, rid, sheet_id, fname, xml_bytes in new_sheet_info:
                zw.writestr(fname, xml_bytes)


def procesar(ruta_archivo1, ruta_archivo2, ruta_salida=None):
    """Proceso principal."""
    print(f"Archivo 1 (Placas): {os.path.basename(ruta_archivo1)}")
    print(f"Archivo 2 (Rutas):  {os.path.basename(ruta_archivo2)}")

    # 1. Extract mapping from Archivo 1
    mapeo_placas = extraer_mapeo_archivo1(ruta_archivo1)
    print(f"\nMapeo de placas extraído ({len(mapeo_placas)} entradas):")
    for pg, pr in sorted(mapeo_placas.items()):
        print(f"  {pg} -> {pr}")

    # 2. Read data from Archivo 2 (read_only = instant)
    wb_ro = openpyxl.load_workbook(ruta_archivo2, read_only=True, data_only=True)
    ws_ruta_ro = wb_ro["RUTA"]
    ws_hoja1_ro = wb_ro["Hoja1"]

    headers = [cell.value for cell in next(ws_ruta_ro.iter_rows(min_row=1, max_row=1))]
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip().upper()] = i

    idx_orden = col_map.get("ORDEN", 0)
    idx_cliente = col_map.get("CLIENTE", 4)
    idx_distrito1 = col_map.get("DISTRITO 1", 8)
    idx_ruta_pc = col_map.get("RUTA PC", 12)
    idx_placa_gen = col_map.get("PLACA GENERICA", 13)
    idx_hora = col_map.get("HORA", 14)
    idx_tipo_und = col_map.get("TIPO UND", 15)

    # Column numbers (1-based) for writing
    col_placa_1b = col_map.get("PLACA", 16) + 1
    col_conductor_1b = col_map.get("CONDUCTOR", 17) + 1

    datos_por_placa_gen = defaultdict(lambda: {
        "orders": [], "clients": set(), "hora": None, "ruta_pc": None,
        "tipo": None, "placa_real": None, "opl": None,
        "por_distrito": defaultdict(lambda: {"orders": [], "clients": set()})
    })

    ruta_cell_updates = {}  # {(row, col): value}

    for row_idx, row in enumerate(ws_ruta_ro.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) <= idx_placa_gen:
            continue
        placa_gen = row[idx_placa_gen]
        if not placa_gen or str(placa_gen).strip() == "":
            continue

        pg_str = str(placa_gen).strip()
        placa_real = mapeo_placas.get(pg_str)
        opl = derivar_opl(pg_str)

        if placa_real:
            ruta_cell_updates[(row_idx, col_placa_1b)] = placa_real
            ruta_cell_updates[(row_idx, col_conductor_1b)] = opl

        orden = row[idx_orden] if len(row) > idx_orden else None
        cliente = row[idx_cliente] if len(row) > idx_cliente else None
        hora = row[idx_hora] if len(row) > idx_hora else None
        ruta_pc = row[idx_ruta_pc] if len(row) > idx_ruta_pc else None
        tipo = row[idx_tipo_und] if len(row) > idx_tipo_und else None
        distrito = row[idx_distrito1] if len(row) > idx_distrito1 else None

        grupo = datos_por_placa_gen[pg_str]
        grupo["orders"].append(orden)
        if cliente:
            grupo["clients"].add(cliente)
        if hora is not None:
            grupo["hora"] = hora
        if ruta_pc:
            grupo["ruta_pc"] = ruta_pc
        if tipo:
            grupo["tipo"] = tipo
        grupo["placa_real"] = placa_real or pg_str
        grupo["opl"] = opl

        if distrito:
            d_grupo = grupo["por_distrito"][str(distrito).strip()]
            d_grupo["orders"].append(orden)
            if cliente:
                d_grupo["clients"].add(cliente)

    # Read Hoja1 data
    hoja1_rows = list(ws_hoja1_ro.iter_rows(values_only=True))
    wb_ro.close()

    print(f"\nDatos procesados: {len(datos_por_placa_gen)} grupos de placa genérica")

    # 3. Prepare Hoja1 modifications
    hoja1_cell_updates = {
        (1, 6): "CONDUCTOR", (1, 7): "All",
        (3, 10): "Recuento de ORDEN", (3, 11): "Recuento distinto de CLIENTE",
    }
    total_clientes = 0
    for r_idx, row_data in enumerate(hoja1_rows[3:], start=4):
        if row_data and len(row_data) > 7:
            placa_gen = row_data[7]  # col H (0-indexed)
            hora_label = row_data[5]  # col F (0-indexed)
            if placa_gen and str(placa_gen).strip() in datos_por_placa_gen:
                n_cli = len(datos_por_placa_gen[str(placa_gen).strip()]["clients"])
                hoja1_cell_updates[(r_idx, 11)] = n_cli
                total_clientes += n_cli
            elif hora_label and "total" in str(hora_label).lower():
                hoja1_cell_updates[(r_idx, 11)] = total_clientes

    # 4. Build Hoja3 (summary by HORA/RUTA PC)
    orden_hoja1 = []
    for row_data in hoja1_rows[3:]:
        if row_data and len(row_data) > 7:
            pg = row_data[7]
            if pg and str(pg).strip() in datos_por_placa_gen:
                orden_hoja1.append(str(pg).strip())

    hoja3_rows = []
    # Header
    h3_headers = ["HORA", "RUTA PC", "PLACA GENERICA", "TIPO UND",
                   "Recuento de ORDEN", "Recuento distinto de CLIENTE", "placa", "OPL"]
    hoja3_rows.append((1, [(c + 1, h, True) for c, h in enumerate(h3_headers)]))

    fila = 2
    hora_actual = None
    total_ordenes = total_clientes = 0
    for pg in orden_hoja1:
        grupo = datos_por_placa_gen[pg]
        hora = grupo["hora"]
        cells = []
        if hora != hora_actual:
            cells.append((1, hora, False))
            hora_actual = hora
        cells.append((2, grupo["ruta_pc"], False))
        cells.append((3, pg, False))
        cells.append((4, grupo["tipo"], False))
        n_ord = len(grupo["orders"])
        n_cli = len(grupo["clients"])
        cells.append((5, n_ord, False))
        cells.append((6, n_cli, False))
        cells.append((7, grupo["placa_real"], False))
        cells.append((8, grupo["opl"], False))
        hoja3_rows.append((fila, cells))
        total_ordenes += n_ord
        total_clientes += n_cli
        fila += 1

    hoja3_rows.append((fila, [
        (1, "Total general", True), (5, total_ordenes, False), (6, total_clientes, False)
    ]))

    hoja3_xml = build_new_sheet_xml(hoja3_rows)

    # 5. Build Hoja8 (pivot by HORA > PLACA > DISTRITO)
    pivot_data = defaultdict(lambda: defaultdict(lambda: defaultdict(
        lambda: {"orders": [], "clients": set()}
    )))
    for pg, grupo in datos_por_placa_gen.items():
        hora = grupo["hora"]
        placa_real = grupo["placa_real"]
        for distrito, d_data in grupo["por_distrito"].items():
            pivot_data[hora][placa_real][distrito]["orders"].extend(d_data["orders"])
            pivot_data[hora][placa_real][distrito]["clients"].update(d_data["clients"])

    def sort_hora(h):
        if h is None:
            return (999, 0)
        if isinstance(h, dt_time):
            return (h.hour, h.minute)
        try:
            return (int(h), 0)
        except (ValueError, TypeError):
            return (999, 0)

    hoja8_rows = []
    hoja8_rows.append((1, [(1, "CONDUCTOR", True), (2, "All", False)]))
    h8_headers = ["HORA", "PLACA", "DISTRITO 1", "Recuento de ORDEN", "Recuento distinto de CLIENTE"]
    hoja8_rows.append((3, [(c + 1, h, True) for c, h in enumerate(h8_headers)]))

    fila = 4
    gran_total_ord = gran_total_cli = 0
    for hora in sorted(pivot_data.keys(), key=sort_hora):
        hora_total_ord = hora_total_cli = 0
        first_in_hora = True
        for placa in sorted(pivot_data[hora].keys()):
            placa_total_ord = placa_total_cli = 0
            first_in_placa = True
            for distrito in sorted(pivot_data[hora][placa].keys()):
                d_data = pivot_data[hora][placa][distrito]
                n_ord = len(d_data["orders"])
                n_cli = len(d_data["clients"])
                cells = []
                if first_in_hora:
                    cells.append((1, hora, False))
                    first_in_hora = False
                if first_in_placa:
                    cells.append((2, placa, False))
                    first_in_placa = False
                cells.append((3, distrito, False))
                cells.append((4, n_ord, False))
                cells.append((5, n_cli, False))
                hoja8_rows.append((fila, cells))
                placa_total_ord += n_ord
                placa_total_cli += n_cli
                fila += 1

            hoja8_rows.append((fila, [
                (2, f"Total {placa}", True), (4, placa_total_ord, False), (5, placa_total_cli, False)
            ]))
            hora_total_ord += placa_total_ord
            hora_total_cli += placa_total_cli
            fila += 1

        hora_str = hora.strftime("%H:%M:%S") if isinstance(hora, dt_time) else str(hora or "")
        hoja8_rows.append((fila, [
            (1, f"Total {hora_str}", True), (4, hora_total_ord, False), (5, hora_total_cli, False)
        ]))
        gran_total_ord += hora_total_ord
        gran_total_cli += hora_total_cli
        fila += 1

    hoja8_rows.append((fila, [
        (1, "Total general", True), (4, gran_total_ord, False), (5, gran_total_cli, False)
    ]))

    hoja8_xml = build_new_sheet_xml(hoja8_rows)

    # 6. Build output file
    if ruta_salida is None:
        base, ext = os.path.splitext(ruta_archivo2)
        ruta_salida = f"{base} - RESULTADO{ext}"

    # First: copy original and modify RUTA + Hoja1 sheets
    tmp_path = ruta_salida + '.tmp'

    with zipfile.ZipFile(ruta_archivo2, 'r') as zr:
        sheet_map = get_sheet_file_map(zr)
        ruta_file = sheet_map.get("RUTA")
        hoja1_file = sheet_map.get("Hoja1")

        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zw:
            for item in zr.infolist():
                data = zr.read(item.filename)
                if item.filename == ruta_file:
                    data = modify_sheet_xml(data, ruta_cell_updates)
                elif item.filename == hoja1_file:
                    data = modify_sheet_xml(data, hoja1_cell_updates)
                zw.writestr(item, data)

    print("  Sheets modified.")

    # Second: add Hoja3 and Hoja8
    add_sheets_to_workbook(tmp_path, ruta_salida, [
        ("Hoja3", hoja3_xml),
        ("Hoja8", hoja8_xml),
    ], insert_before_sheet="RUTA")

    os.remove(tmp_path)
    print(f"\nArchivo guardado: {ruta_salida}")
    return ruta_salida


def main():
    if len(sys.argv) == 3:
        archivo1, archivo2 = sys.argv[1], sys.argv[2]
    elif len(sys.argv) == 2:
        archivo1, archivo2 = detectar_archivos(sys.argv[1])
    else:
        archivo1, archivo2 = detectar_archivos(".")
    print("=" * 60)
    print("SISTEMA DE PROCESAMIENTO EXCEL - PICKING CENTER")
    print("=" * 60)
    procesar(archivo1, archivo2)
    print("\n¡Proceso completado exitosamente!")


if __name__ == "__main__":
    main()
